"""Dựng báo cáo thu/chi theo tháng ra file Excel (.xlsx) + lưu siêu dữ liệu.

Báo cáo CHỈ tổng hợp lại dữ liệu kế toán đã được người duyệt trong bảng
`transactions` (deleted_at IS NULL) — không đụng tới luồng OCR. Mỗi tháng 1 bản
(period 'YYYY-MM'); tạo lại sẽ ghi đè file + dòng cùng tháng.

openpyxl là thuần Python (không cần native deps) nên giữ backend nhẹ đúng triết
lý dự án.
"""
from __future__ import annotations

import calendar
import re
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .audit import log_activity
from .config import get_settings
from .models import UserOut
from .notify import notify_admins
from .timezone import local_today

PERIOD_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")

# Màu lấy từ COLORS.md (bỏ dấu '#' theo định dạng openpyxl ARGB/RGB).
_BRAND = "2563EB"        # brand-600 — tiêu đề, header bảng
_BRAND_DARK = "1D4ED8"   # brand-700
_INK = "131A27"          # ink-900
_EMERALD = "059669"      # thu
_ROSE = "E11D48"         # chi
_SLATE_900 = "0F172A"
_SLATE_700 = "334155"
_SLATE_500 = "64748B"
_SLATE_200 = "E2E8F0"
_SLATE_50 = "F8FAFC"
_BRAND_50 = "EFF6FF"
_EMERALD_50 = "ECFDF5"
_ROSE_50 = "FFF1F2"
_WHITE = "FFFFFF"

_MONEY_FMT = '#,##0" ₫"'
_VN_MONTHS = "tháng {month}/{year}"

_thin = Side(style="thin", color=_SLATE_200)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# --------------------------------------------------------------------------
# Tiện ích kỳ báo cáo
# --------------------------------------------------------------------------
def period_range(period: str) -> tuple[str, str]:
    """'YYYY-MM' -> ('YYYY-MM-01', 'YYYY-MM-<ngày cuối>')."""
    year, month = int(period[:4]), int(period[5:7])
    last = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last:02d}"


def previous_period() -> str:
    """Tháng đã chốt gần nhất (tháng trước theo giờ địa phương)."""
    today = local_today()
    year, month = today.year, today.month
    if month == 1:
        return f"{year - 1:04d}-12"
    return f"{year:04d}-{month - 1:02d}"


def current_period() -> str:
    today = local_today()
    return f"{today.year:04d}-{today.month:02d}"


def _vn_date(iso: str) -> str:
    """'YYYY-MM-DD' -> 'DD/MM/YYYY'."""
    if not iso or len(iso) < 10:
        return iso
    return f"{iso[8:10]}/{iso[5:7]}/{iso[0:4]}"


def report_filename(period: str) -> str:
    """Tên file tải về thân thiện: bao_cao_thu_chi_2026-05.xlsx."""
    return f"bao_cao_thu_chi_{period}.xlsx"


# --------------------------------------------------------------------------
# Lấy dữ liệu
# --------------------------------------------------------------------------
def _gather(conn: sqlite3.Connection, period: str) -> dict:
    start, end = period_range(period)
    summary = conn.execute(
        "SELECT "
        "COALESCE(SUM(CASE WHEN kind='income' THEN amount END),0) AS income, "
        "COALESCE(SUM(CASE WHEN kind='expense' THEN amount END),0) AS expense, "
        "COUNT(*) AS cnt "
        "FROM transactions WHERE deleted_at IS NULL AND txn_date BETWEEN ? AND ?",
        (start, end),
    ).fetchone()
    daily = conn.execute(
        "SELECT txn_date AS d, "
        "COALESCE(SUM(CASE WHEN kind='income' THEN amount END),0) AS income, "
        "COALESCE(SUM(CASE WHEN kind='expense' THEN amount END),0) AS expense "
        "FROM transactions WHERE deleted_at IS NULL AND txn_date BETWEEN ? AND ? "
        "GROUP BY txn_date ORDER BY txn_date",
        (start, end),
    ).fetchall()
    txns = conn.execute(
        "SELECT txn_date, room, note, kind, amount FROM transactions "
        "WHERE deleted_at IS NULL AND txn_date BETWEEN ? AND ? "
        "ORDER BY txn_date, id",
        (start, end),
    ).fetchall()
    income, expense = summary["income"], summary["expense"]
    return {
        "start": start,
        "end": end,
        "total_income": income,
        "total_expense": expense,
        "balance": income - expense,
        "txn_count": summary["cnt"],
        "daily": daily,
        "txns": txns,
    }


# --------------------------------------------------------------------------
# Dựng workbook Excel
# --------------------------------------------------------------------------
def _title_font(size: int, color: str = _SLATE_900) -> Font:
    return Font(name="Calibri", size=size, bold=True, color=color)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_BRAND)


def _set_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_summary_sheet(ws, period: str, data: dict, hotel: str) -> None:
    ws.title = "Tổng hợp"
    ws.sheet_view.showGridLines = False
    _set_widths(ws, {1: 26, 2: 20, 3: 20, 4: 20})
    year, month = int(period[:4]), int(period[5:7])

    # Tiêu đề
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = hotel
    c.font = _title_font(18, _INK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"BÁO CÁO THU CHI {_VN_MONTHS.format(month=month, year=year).upper()}"
    c.font = _title_font(14, _BRAND)
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:D3")
    c = ws["A3"]
    c.value = (
        f"Kỳ báo cáo: {_vn_date(data['start'])} – {_vn_date(data['end'])}   |   "
        f"Lập ngày: {_vn_date(local_today().isoformat())}"
    )
    c.font = Font(name="Calibri", size=10, italic=True, color=_SLATE_500)

    # Thẻ KPI (hàng nhãn + hàng giá trị)
    kpis = [
        ("TỔNG THU", data["total_income"], _EMERALD, _EMERALD_50),
        ("TỔNG CHI", data["total_expense"], _ROSE, _ROSE_50),
        ("TỒN QUỸ (THU - CHI)", data["balance"], _BRAND, _BRAND_50),
        ("SỐ CHỨNG TỪ", data["txn_count"], _SLATE_700, _SLATE_50),
    ]
    label_row, value_row = 5, 6
    ws.row_dimensions[label_row].height = 18
    ws.row_dimensions[value_row].height = 26
    for i, (label, value, color, bg) in enumerate(kpis, start=1):
        lc = ws.cell(row=label_row, column=i, value=label)
        lc.font = Font(name="Calibri", size=9, bold=True, color=_SLATE_500)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.border = _BORDER
        vc = ws.cell(row=value_row, column=i, value=value)
        vc.font = Font(name="Calibri", size=14, bold=True, color=color)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.border = _BORDER
        if label != "SỐ CHỨNG TỪ":
            vc.number_format = _MONEY_FMT

    # Bảng theo ngày
    head_row = 9
    ws.cell(row=head_row - 1, column=1, value="CHI TIẾT THEO NGÀY").font = _title_font(11, _SLATE_700)
    headers = ["Ngày", "Thu", "Chi", "Chênh lệch"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=head_row, column=i, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=_WHITE)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[head_row].height = 20

    r = head_row + 1
    for day in data["daily"]:
        inc, exp = day["income"], day["expense"]
        values = [_vn_date(day["d"]), inc, exp, inc - exp]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = _BORDER
            cell.font = Font(name="Calibri", size=10, color=_SLATE_700)
            if i == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.number_format = _MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
        r += 1

    # Hàng tổng cộng
    total_vals = ["TỔNG CỘNG", data["total_income"], data["total_expense"], data["balance"]]
    for i, v in enumerate(total_vals, start=1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.fill = PatternFill("solid", fgColor=_SLATE_50)
        cell.font = Font(name="Calibri", size=10, bold=True, color=_SLATE_900)
        cell.border = _BORDER
        if i == 1:
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.number_format = _MONEY_FMT
            cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A10"


def _build_detail_sheet(ws, data: dict) -> None:
    ws.title = "Chi tiết"
    ws.sheet_view.showGridLines = False
    _set_widths(ws, {1: 6, 2: 14, 3: 24, 4: 40, 5: 10, 6: 18})

    headers = ["STT", "Ngày", "Phòng / Khách", "Nội dung", "Loại", "Số tiền"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=_WHITE)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[1].height = 20

    r = 2
    for idx, t in enumerate(data["txns"], start=1):
        is_income = t["kind"] == "income"
        kind_label = "Thu" if is_income else "Chi"
        money_color = _EMERALD if is_income else _ROSE
        cells = [
            (idx, "center", None, _SLATE_700),
            (_vn_date(t["txn_date"]), "center", None, _SLATE_700),
            (t["room"], "left", None, _SLATE_700),
            (t["note"], "left", None, _SLATE_700),
            (kind_label, "center", None, money_color),
            (t["amount"], "right", _MONEY_FMT, money_color),
        ]
        for i, (v, align, fmt, color) in enumerate(cells, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = _BORDER
            bold = i in (5, 6)
            cell.font = Font(name="Calibri", size=10, bold=bold, color=color)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(i == 4))
            if fmt:
                cell.number_format = fmt
        r += 1

    if not data["txns"]:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        c = ws.cell(row=2, column=1, value="Không có chứng từ nào trong kỳ.")
        c.font = Font(name="Calibri", size=10, italic=True, color=_SLATE_500)
        c.alignment = Alignment(horizontal="center")
        r = 3

    # Tổng cộng
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    label = ws.cell(row=r, column=1, value="TỔNG THU - CHI")
    label.font = Font(name="Calibri", size=10, bold=True, color=_SLATE_900)
    label.alignment = Alignment(horizontal="right", vertical="center")
    label.fill = PatternFill("solid", fgColor=_SLATE_50)
    total = ws.cell(row=r, column=6, value=data["balance"])
    total.number_format = _MONEY_FMT
    total.font = Font(name="Calibri", size=10, bold=True, color=_SLATE_900)
    total.alignment = Alignment(horizontal="right", vertical="center")
    total.fill = PatternFill("solid", fgColor=_SLATE_50)
    total.border = _BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{max(1, r - 1)}"


def build_workbook(period: str, data: dict, hotel: str) -> Workbook:
    wb = Workbook()
    _build_summary_sheet(wb.active, period, data, hotel)
    _build_detail_sheet(wb.create_sheet(), data)
    return wb


# --------------------------------------------------------------------------
# Tạo + lưu báo cáo (KHÔNG commit — để caller commit cùng giao dịch)
# --------------------------------------------------------------------------
def generate_report(
    conn: sqlite3.Connection,
    period: str,
    *,
    user: UserOut | None = None,
    auto: bool = False,
) -> int:
    """Render Excel, ghi file ra đĩa, upsert dòng `reports`, log + notify.

    Trả về id của dòng báo cáo. Caller chịu trách nhiệm commit.
    """
    if not PERIOD_RE.fullmatch(period):
        raise ValueError("Kỳ báo cáo phải có dạng YYYY-MM")

    settings = get_settings()
    data = _gather(conn, period)
    wb = build_workbook(period, data, settings.hotel_name)

    settings.report_path.mkdir(parents=True, exist_ok=True)
    dest = settings.report_path / report_filename(period)
    wb.save(dest)

    title = f"Báo cáo thu chi {_VN_MONTHS.format(month=int(period[5:7]), year=int(period[:4]))}"
    # Upsert theo period (UNIQUE) — tạo lại sẽ cập nhật bản cũ thay vì thêm mới.
    conn.execute(
        "INSERT INTO reports "
        "(period, title, file_path, total_income, total_expense, balance, txn_count, auto, generated_by) "
        "VALUES (?,?,?,?,?,?,?,?,?) "
        "ON CONFLICT(period) DO UPDATE SET "
        "title=excluded.title, file_path=excluded.file_path, "
        "total_income=excluded.total_income, total_expense=excluded.total_expense, "
        "balance=excluded.balance, txn_count=excluded.txn_count, "
        "auto=excluded.auto, generated_by=excluded.generated_by, "
        "updated_at=datetime('now')",
        (
            period, title, str(dest), data["total_income"], data["total_expense"],
            data["balance"], data["txn_count"], 1 if auto else 0,
            user.id if user else None,
        ),
    )
    row = conn.execute("SELECT id FROM reports WHERE period = ?", (period,)).fetchone()
    report_id = row["id"]

    log_activity(
        conn, user, "report.generate", target_type="report", target_id=report_id,
        detail={"period": period, "auto": auto, "txn_count": data["txn_count"]},
    )
    notify_admins(
        conn,
        type="report.ready",
        level="success",
        title="Báo cáo tháng đã sẵn sàng",
        body=(
            f"{title} đã được hệ thống tự tạo." if auto
            else f"{(user.full_name or user.username) if user else 'Người dùng'} vừa tạo {title.lower()}."
        ),
        link="/reports",
        actor=user,
        target_type="report",
        target_id=report_id,
        event_key_prefix=f"report.ready:{period}" if auto else None,
    )
    return report_id
